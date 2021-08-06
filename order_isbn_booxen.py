from tkinter import *
import tkinter.font as ft
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time , datetime

def main():
    main = Tk()
    main.title("북쎈 ISBN 봇")
    main.geometry("400x200")
    
    global ID
    global PW
    ID = "naerok2"
    PW = "ipass3309="

    button_target = Button(main, text="시작", command=lambda : macro(), fg="black", width=5, height =2).place(x = 300, y = 18)
    main.mainloop()

def macro():
    url = "https://orderbook.booxen.com/"
    driver = webdriver.Firefox(executable_path=r'D:\Auto Bak\Desktop\macro\geckodriver.exe')
    driver.maximize_window()
    driver.get(url)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="userId"]').send_keys(ID)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="pass"]').send_keys(PW)
    driver.find_element_by_xpath('//*[@id="adm-intro"]/div[1]/div[1]/div/div[2]/div[2]/div/form/button/img').click()
    time.sleep(6)
    searchInfo(driver)

def searchInfo(driver):
    global count
    count = 2
    deleteNo = 0
    bookList=load_workbook(filename='test.xlsx')
    sheet = bookList.active
    max = sheet.max_row
    bookList.close()
    for i in range(max-1):
        try:
            bookList=load_workbook(filename='test.xlsx')
            sheet = bookList.active
            book = sheet.cell(count,2).value
            WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="kwdGnb"]'))).send_keys(book)# find_element_by_xpath('//*[@id="kwdGnb"]').send_keys(book)
            WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="selForm"]/fieldset/div[2]/a/img'))).click()
  
            per = WebDriverWait(driver, 5).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="itemPrice0"]')))[1].text
            per = per[-4:]
            per = per.replace(')', '')

            price = driver.find_elements_by_xpath('//*[@id="itemPrice0"]')[1].text
            price = price[5:12]
            price = price.replace('원', '')
            price = price.replace(',', '')
            try:
                price = price.replace(' ', '')
            except:
                pass
            
            ownPrice = sheet.cell(count, 9).value
            sheet.cell(count, 20, "{}".format(per))
            bookList.save("test.xlsx")
            bookList.close()

            okPrice = int(price) + int(ownPrice/10)

            okToSell(count, per, ownPrice, okPrice, deleteNo)
        except:
            sheet.cell(count, 20, "100%")
            bookList.save("test.xlsx")
            bookList.close()
        finally:
            betterOne(count)
            driver.find_element_by_xpath('//*[@id="kwdGnb"]').clear()
        count+=1

def betterOne(count):
    bookList = load_workbook(filename = "test.xlsx")
    sheet = bookList.active
    kyobo = sheet.cell(count, 19).value
    boosen = sheet.cell(count, 20).value

    kyobo = kyobo.replace("%", "")
    boosen = boosen.replace("%", "")
    print(sheet.cell(count, 10).value)
    kyobo = int(kyobo)
    boosen = int(boosen)

    if(kyobo == 100 and boosen == 100):
        sheet.cell(count, 1, "★")
        bookList.save("test.xlsx")
        bookList.close()
    elif(kyobo >= boosen):
        sheet.cell(count, 1, "북쎈")
        bookList.save("test.xlsx")
        bookList.close()
    elif(kyobo < boosen):
        sheet.cell(count, 1, "교보")
        bookList.save("test.xlsx")
        bookList.close()
    else:
        sheet.cell(count, 1, "★")
        bookList.save("test.xlsx")
        bookList.close()
    

def okToSell(count, per, ownPrice, okPrice, deleteNo):
    print("own : {} , ok : {}, per : {}, count : {}, deleteNo : {}".format(ownPrice, okPrice, per, count, deleteNo))
    if(ownPrice < okPrice):
        bookList=load_workbook(filename='test.xlsx')
        sheet = bookList.active
        sheet.cell(count, 20, "100%")
        bookList.save("test.xlsx")
        bookList.close()
    else:
        bookList=load_workbook(filename='test.xlsx')
        sheet = bookList.active
        sheet.cell(count, 20, per)
            
main()