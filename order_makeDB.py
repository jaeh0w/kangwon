from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox


def main():
    root = Tk()
    root.geometry("350x180")
    Label_raw_fileName = Label(root, text="RAW 액셀파일 이름 :").place(x = 20, y = 20)
    Entry_raw_fileName = Entry(root)
    Entry_raw_fileName.place(x = 150, y =  20)

    Label_final_fileName = Label(root, text="가공된 액셀파일 이름 :").place(x = 12, y = 50)
    Entry_final_fileName = Entry(root)
    Entry_final_fileName.place(x = 150, y =  50)

    Label_pub = Label(root, text="거래처 이름 :").place(x = 61, y = 80)
    Entry_pub = Entry(root)
    Entry_pub.place(x = 150, y =  80)
    root.title("입고 프로그램")

    Button_start = Button(root, text="프로그램 가동", command=lambda : writeExcel(Entry_raw_fileName.get(), Entry_pub.get(), Entry_final_fileName.get()), width=25, height=2).place(x = 35, y = 120)
    root.mainloop()
    bookList = load_workbook()

def writeExcel(fileName, partner, final_fileName):
    if(partner == "교보"):
        partner = "(주)교보문고"
        kyobo(fileName, partner, final_fileName)
    elif(partner == "북쎈"):
        partner = "주식회사웅진북쎈"
        booxen(fileName, partner, final_fileName)
    else:
        messagebox.showinfo("알림", "올바른 출판사를 입력하세요!\n교보 / 북쎈")

def kyobo(fileName, partner, final_fileName):
    goExcel_name = "{}.xlsx".format(fileName)
    goExcel = load_workbook(goExcel_name)
    goExcel_ws = goExcel.active
    max = goExcel_ws.max_row
    goExcel.close()

    mainExcel_name = "{}.xlsx".format(final_fileName)
    mainExcel = load_workbook(mainExcel_name)
    mainExcel_ws = mainExcel.active
    first = mainExcel_ws.max_row
    print(first)
    for i in range(3, max):
        when = str(goExcel_ws.cell(i, 2).value)
        isbn = goExcel_ws.cell(i, 5).value
        name = goExcel_ws.cell(i, 6).value
        pub = goExcel_ws.cell(i, 7).value
        auth = goExcel_ws.cell(i, 8).value
        ownPrice = goExcel_ws.cell(i, 10).value
        per = goExcel_ws.cell(i, 12).value
        amount = goExcel_ws.cell(i, 11).value
        price = goExcel_ws.cell(i, 14).value
        when = when.replace("-", ".")
        
        inPrice = price.replace(",", "")

        IN = int((int(isbn)*int(int(inPrice)+i))/(10000000-i))

        mainExcel_ws.cell(first, 1, "IN{}".format(IN))
        mainExcel_ws.cell(first, 2, when)
        mainExcel_ws.cell(first, 3, partner)
        mainExcel_ws.cell(first, 4, isbn)
        mainExcel_ws.cell(first, 5, name)
        mainExcel_ws.cell(first, 6, pub)
        mainExcel_ws.cell(first, 7, auth)
        mainExcel_ws.cell(first, 8, ownPrice)
        mainExcel_ws.cell(first, 9, per)
        mainExcel_ws.cell(first, 10, amount)
        mainExcel_ws.cell(first, 11, price)
        value = ["{}".format(IN), when, partner, isbn, name, pub, ownPrice, per, amount, price]
        print(value)
        first+=1
    first = 0
    mainExcel.save(mainExcel_name)

def booxen(fileName, partner, final_fileName):
    goExcel_name = "{}.xlsx".format(fileName)
    goExcel = load_workbook(goExcel_name)
    goExcel_ws = goExcel.active
    max = goExcel_ws.max_row

    goExcel.close()

    mainExcel_name = "{}.xlsx".format(final_fileName)
    mainExcel = load_workbook(mainExcel_name)
    mainExcel_ws = mainExcel.active
    first = mainExcel_ws.max_row
    for i in range(3, max):
        when = str(goExcel_ws.cell(i, 3).value)
        isbn = goExcel_ws.cell(i, 12).value
        name = goExcel_ws.cell(i, 14).value
        pub = goExcel_ws.cell(i, 11).value
        ownPrice = goExcel_ws.cell(i, 16).value
        per = goExcel_ws.cell(i, 17).value
        amount = goExcel_ws.cell(i, 21).value
        price = goExcel_ws.cell(i, 23).value

        per = "{}%".format(per)
        when = when[0:10]
        when = when.replace("-", ".")
        IN = int((int(isbn)*int(price+i))/(10000000-i))
        price = "{:,}".format(price)
        ownPrice = "{:,}".format(ownPrice)
        #mainExcel_ws.append(value)
        mainExcel_ws.cell(first, 1, "IN{}".format(IN))
        mainExcel_ws.cell(first, 2, when)
        mainExcel_ws.cell(first, 3, partner)
        mainExcel_ws.cell(first, 4, isbn)
        mainExcel_ws.cell(first, 5, name)
        mainExcel_ws.cell(first, 6, pub)
        mainExcel_ws.cell(first, 7, "auth")
        mainExcel_ws.cell(first, 8, ownPrice)
        mainExcel_ws.cell(first, 9, per)
        mainExcel_ws.cell(first, 10, amount)
        mainExcel_ws.cell(first, 11, price)
        value = ["{}".format(IN), when, partner, isbn, name, pub, ownPrice, per, amount, price]
        print(value)
        first+=1
        
    mainExcel.save(mainExcel_name)
    

main()