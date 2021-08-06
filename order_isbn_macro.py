from tkinter import *
import tkinter.font as ft
from typing import final
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time , datetime

def main():
    main = Tk()
    main.title("ISBN 봇")
    main.geometry("400x200")

    global ID
    global PW
    ID = "1544200122"
    PW = "ipass2255-"
    button_target = Button(main, text="시작", command=lambda : macro(), fg="green", width=5, height =2).place(x = 300, y = 18)
    main.mainloop()

def macro():
    baseUrl = "https://bscm.kyobobook.co.kr"
    # url = baseUrl+str((plusUrl))
    url = baseUrl
    driver = webdriver.Firefox(executable_path=r'D:\Auto Bak\Desktop\macro\geckodriver.exe')
    driver.maximize_window()
    driver.get(url)
    time.sleep(1)

    driver.find_element_by_class_name("w2input.inp_id").send_keys(ID)
    time.sleep(1)
    driver.find_element_by_class_name("w2input.inp_pw").send_keys(PW)
    driver.find_element_by_class_name("w2anchor2.anc_login_pc").click()
    
    time.sleep(10)
    searchInfo(driver)

def searchInfo(driver):
    bookList=load_workbook(filename='test.xlsx')
    sheet = bookList.active
    count = 2
    deleteNo = 0
    max = sheet.max_row
    bookList.close()
    for i in range(max-1):
        try:
            bookList=load_workbook(filename='test.xlsx')
            sheet = bookList.active
            book = sheet.cell(count,2).value
            WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="mf_wfm_header_ibx_findName"]'))).send_keys(book)
            WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.CLASS_NAME, 'w2textbox.search_tit'))).click()
            isbn = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_0_tbx_cmdtCode"]'))).text 
            per = driver.find_element_by_xpath('//*[@id="mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_0_tbx_byngRate"]').text
            per = per.replace('(', '')
            per = per.replace(')', '')
            price = driver.find_element_by_xpath('//*[@id="mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_0_tbx_byngPrce"]').text
            price = price.replace(',', '')
            price = price.replace('원', '')
            sheet.cell(count, 19, "{}".format(per))
            bookList.save("test.xlsx")
            bookList.close()
            okToSell(count, per, price, isbn)
        except:
            sheet.cell(count, 19, "100%")
            bookList.save("test.xlsx")
            bookList.close()
        count +=1

def okToSell(count, per, price,deleteNo):
    bookList=load_workbook(filename='test.xlsx')
    sheet = bookList.active
    ownPrice = sheet.cell(count,9).value
    bookList.close()
    okPrice = int(price) + int(ownPrice/10)
    print(ownPrice)
    print(okPrice)

    if(per == "85%"):
        print("X")
        bookList=load_workbook(filename='test.xlsx')
        sheet = bookList.active
        sheet.cell(count, 19, "100%")
        bookList.save("test.xlsx")
        bookList.close()

    elif(ownPrice < okPrice):
        print("X")
        bookList=load_workbook(filename='test.xlsx')
        sheet = bookList.active
        sheet.cell(count, 19, "100%")
        bookList.save("test.xlsx")
        bookList.close()
    else:
        print("O")
        wb = load_workbook(filename="test.xlsx")
        ws = wb.active
        ws.cell(count, 19, per)
        wb.save("test.xlsx")
        wb.close()
main()