from os import lstat, pwrite
from tkinter import *
import tkinter.font as ft
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time , datetime



def main():
    main = Tk()
    main.title("ISBN 봇")
    main.geometry("400x200")

    myFont = ft.Font(family = "나눔고딕")
    button_target = Button(main, text="시작", command=lambda : macro(), fg="green", width=5).place(x = 300, y = 18)
    main.mainloop()

def macro():
    url = "https://bscm.kyobobook.co.kr"
    global ID
    global PW
    ID = "1544200122"
    PW = "ipass2255-"
    driver = webdriver.Safari()
    driver.maximize_window()
    driver.get(url)
    time.sleep(1)

    driver.find_element_by_class_name("w2input.inp_id").send_keys(ID)
    driver.find_element_by_class_name("w2input.inp_pw").send_keys(PW)
    driver.find_element_by_class_name("w2anchor2.anc_login_pc").click()
    
    time.sleep(9)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="mf_wfm_header_ibx_findName"]'))).text
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="mf_wfm_header_ibx_findName"]').send_keys("소명출판")
    time.sleep(0.5)
    driver.find_element_by_class_name('w2textbox.search_tit').click()
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="mf_wfm_content_tac_main_contents_content1_body_btn_date"]').click()

    time.sleep(3)
    searchInfo(driver)


def searchInfo(driver):
    count = 1
    time.sleep(1)
    lst_title = []
    lst_isbn = []
    lst_publisher = []
    lst_auth = []
    lst_date = []
    lst_corPrice = []
    lst_okPrice = []
    lst_rate = []
    for i in range(2, 101):
        for j in range(20):
            lst_okPrice.append(driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_"+ str(j) + "_tbx_byngPrce").text)
            lst_title.append(driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_"+ str(j) + "_tbx_cmdtName").text)
            lst_isbn.append(driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_"+ str(j) + "_tbx_cmdtCode").text)
            lst_publisher.append(driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_"+ str(j) +"_tbx_pbcmName").text)
            lst_publisher.append(driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_"+ str(j) +"_tbx_pbcmName").text)
            lst_auth.append(driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_"+ str(j) +"_tbx_autrName").text)
            lst_date.append(driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_"+ str(j) +"_tbx_rlseDate").text)
            lst_corPrice.append(driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_"+ str(j) +"_tbx_wncrPrce").text)
            lst_rate.append(driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_gen_bksSrch_"+ str(j) +"_tbx_byngRate").text)
        print(lst_title)
        for a in range(20):
            writeExcel(lst_title[a], lst_isbn[a], lst_publisher[a], lst_auth[a],lst_date[a], lst_corPrice[a], lst_okPrice[a], lst_rate[a], count )
            count+=1
        time.sleep(1.5)
        if(i % 10 == 1):
            driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_pgc_top_next_btn").click()
        else:
            driver.find_element_by_id("mf_wfm_content_tac_main_contents_content1_body_pgc_top_page_"+str(i)).click()
        time.sleep(3)

def writeExcel(title, isbn, publisher,auth, date, corPrice, okPrice, rate, count):
    bookList=load_workbook(filename='pub.xlsx')
    sheet = bookList.active
    sheet.cell(count,1,title)
    sheet.cell(count,2,isbn)
    sheet.cell(count,3,publisher)
    sheet.cell(count,4,auth)
    sheet.cell(count,5,date)
    sheet.cell(count,6,corPrice)
    sheet.cell(count,7,okPrice)
    sheet.cell(count,8,rate)
    bookList.save("pub.xlsx")


main()