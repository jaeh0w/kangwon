from tkinter import *
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
import time , datetime
import pandas as pd

def main():
    main = Tk()
    main.title("YES24 ISBN 봇")
    main.geometry("400x200")

    button_target = Button(main, text="시작", command=lambda : init(1), fg="green", bg = "black", width=30, height=4).place(x = 40, y = 100)
    main.mainloop()

def init(t):
    url = "http://www.yes24.com/main/default.aspx"
    driver = webdriver.Firefox(executable_path=r'D:\Auto Bak\Desktop\macro\geckodriver.exe')
    driver.maximize_window()
    driver.get(url)
    excel_name = "test.xlsx"
    wb = load_workbook(filename='test.xlsx')
    ws = wb.active
    sheet = pd.read_excel("test.xlsx")
    sheet = sheet.drop_duplicates(['name'])
    sheet.to_excel("test.xlsx", sheet_name='sheet1', index = False, header=True)
    count = 2
    deleteNo = 0
    max = ws.max_row
    wb.close()
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="divYes24SCMEvent"]/div[2]/div[1]/label'))).click()
    for i in range(1, max):
        try:
            wb = load_workbook(filename='test.xlsx')
            ws = wb.active
            bookName = ws.cell(count,10).value
            publisher = ws.cell(count, 17).value
            ownPrice = ws.cell(count, 18).value
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="yesSForm"]/fieldset/span[1]'))).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="query"]'))).clear()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="query"]'))).send_keys(bookName + " "+ publisher)
            time.sleep(1)
            driver.find_element_by_class_name('schBtn').click()
            WebDriverWait(driver, 7).until(EC.presence_of_element_located((By.CLASS_NAME, 'img_bdr'))).click()
            price = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[4]/div[2]/div[2]/div[1]/div[1]/table/tbody/tr[1]/td/span/em'))).text
            price = price.replace(',', '')
            price = price.replace('원', '')
            isbn = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="infoset_specific"]/div[2]/div/table/tbody/tr[3]/td'))).text
            wb.close()
            checkPrice(ownPrice, price, count, deleteNo, isbn)
        except:
            ws.cell(count,1,"★")
            ws.cell(count,2,"★")
            wb.save("test.xlsx")
            wb.close()
        count+=1

def checkPrice(ownPrice, price, count, deleteNo, isbn):
    wb = load_workbook(filename='test.xlsx')
    ws = wb.active
    if(int(ownPrice) == int(price)):
        print("==")
        ws.cell(count,2,"{}".format(int(isbn)))
        wb.save("test.xlsx")
        wb.close()
    else:
        print("X")
        ws.cell(count, 1, "★")
        ws.cell(count, 2, "★")
        wb.save("test.xlsx")
        wb.close()

main()