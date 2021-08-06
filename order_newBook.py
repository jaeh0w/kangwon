from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import tkinter.font as ft
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import pandas as pd


def main():
    root = Tk()
    root.geometry("350x210")

    Label_raw_fileName = Label(root, text="RAW 액셀파일 이름 :").place(x = 20, y = 20)
    Entry_raw_fileName = Entry(root)
    Entry_raw_fileName.place(x = 150, y =  20)

    Label_mid_fileName = Label(root, text="중간처리 액셀파일 이름 :").place(x = 5, y = 50)
    Entry_mid_fileName = Entry(root)
    Entry_mid_fileName.place(x = 150, y =  50)

    Label_final_fileName = Label(root, text="가공된 액셀파일 이름 :").place(x = 17, y = 80)
    Entry_final_fileName = Entry(root)
    Entry_final_fileName.place(x = 150, y =  80)
        
    Label_del_pub = Label(root, text= "삭제할 출판사명 :").place(x = 43, y = 110)
    Entry_del_pub = Entry(root)
    Entry_del_pub.place(x = 150, y =  110)

    root.title("신간 정리 프로그램")

    Button_start = Button(root, text="프로그램 가동", command=lambda : make_one_file(Entry_raw_fileName.get(), Entry_mid_fileName.get(),Entry_final_fileName.get(), Entry_del_pub.get()), width=25, height=2).place(x = 35, y = 150)
    root.mainloop()

def make_one_file(rawFile, midFile,finalFile, pub):
    df = pd.concat(pd.read_excel('{}.xlsx'.format(rawFile), sheet_name=None), ignore_index=True)
    df = df.dropna(axis=0)

    delete_row = df[df["정가"]=="정가"].index
    df = df.drop(delete_row)

    delete_row = df[df["출판사명"]==pub].index
    df = df.drop(delete_row)

    df.to_excel('{}.xlsx'.format(midFile), header=True)

    wb = load_workbook("{}.xlsx".format(midFile))
    ws = wb.active

    for i in range(2, ws.max_row):
        bookName = ws.cell(i, 2).value
        kdc = ws.cell(i, 3).value
        auth = ws.cell(i, 4).value
        pageType = ws.cell(i, 5).value
        pageAmount = ws.cell(i, 6).value
        price = ws.cell(i, 7).value
        publisher = ws.cell(i, 8).value
        isbn = ws.cell(i, 9).value

        print("bookName = {}, kdc = {}, auth = {}, price = {}, pub = {}, isbn = {}".format(bookName, kdc, auth, price, pub, isbn))
        print()

    print(df)

main()