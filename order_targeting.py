from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import tkinter.font as ft
from openpyxl.descriptors.base import Integer
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
import time , datetime

def main():
    main = Tk()
    main.title("타겟팅 봇")
    main.geometry("400x200")

    myFont = ft.Font(family = "나눔고딕")
    label_target = Label(main, text="타겟 대상 : ", font=myFont).place(x = 40, y = 20)
    label_excel = Label(main, text="엑셀 이름 : ", font=myFont).place(x = 40, y = 50)
    label_max = Label(main, text="최대 페이지 : ", font=myFont).place(x = 28, y = 80)
    entry_target = Entry(main)
    entry_target.place(x = 130, y= 20)

    global entry_starting
    entry_starting = Entry(main)
    entry_starting.place(x = 130, y = 110)
    lalbel_starting = Label(main, text = "시작 페이지 : ", font = myFont).place(x = 28, y = 110)

    global entry_excel 
    entry_excel= Entry(main)
    entry_excel.place(x = 130, y= 50)
    global entry_max 
    entry_max = Entry(main)
    entry_max.place(x = 130, y= 80)
    button_target = Button(main, text="시작", command=lambda : targeting(entry_target.get()), fg="white", bg = "green", width=30, height=2).place(x = 40, y = 140)
    main.mainloop()


def targeting(t):
    if(str(t) == "북코치"):
        baseUrl = "http://www.yes24.com//24/usedShop/mall/cjudeer/main"
    elif(str(t) == "마음북"):
        baseUrl = "http://www.yes24.com//24/usedShop/mall/saguraba/main"
    elif(str(t) == "할인에할인"):
        baseUrl = "http://www.yes24.com//24/usedShop/mall/freeman001/main"
    elif(str(t) == "굿북스"):
        baseUrl = "http://www.yes24.com//24/usedShop/mall/ybe7815/main"
    else:
        messagebox.showinfo("알림", "올바른 대상을 입력하세요!")
    #http://www.yes24.com//24/usedShop/mall/saguraba/main 마음북
    #http://www.yes24.com//24/usedShop/mall/cjudeer/main 북코치
    #http://www.yes24.com//24/usedShop/mall/freeman001/main 할인에할인
    # url = baseUrl+str((plusUrl))
    url = baseUrl
    driver = webdriver.Firefox(executable_path=r'D:\Auto Bak\Desktop\macro\geckodriver')
    driver.maximize_window()
    driver.get(url)
    time.sleep(1)

    driver.find_element_by_xpath('//*[@id="divYes24SCMEvent"]/div[2]/div[1]/label').click()

    getInfo(driver, t)


def getInfo(driver, t):
    excel_name = str(entry_excel.get()) + ".xlsx"
    wb = load_workbook(excel_name)
    ws = wb.active
    count = ws.max_row
    wb.close()
    xNo = 1
    max = int(entry_max.get())
    start = int(entry_starting.get())
    if(str(t) == "북코치"):
        seller = "cjudeer"
        sellerID = "242444"
    elif(str(t) == "마음북"):
        seller = "saguraba"
        sellerID = "242487"
    elif(str(t) == "할인에할인"):
        seller = "freeman001"
        sellerID = "235664"
    elif(str(t) == "굿북스"):
        seller = "ybe7815"
        sellerID = "235829"
    if(start == 1):
        xNo = 2
    if(start-10 > 0):
        for m in range(int(start/10)):
            href = "#entrno={}&mallid={}&searchname=&filtername=NAME&PageNumber=".format(str(sellerID),str(seller)) + str(xNo)
            #entrno=235664&mallid=freeman001&searchname=&filtername=NAME&PageNumber=2
            #time.sleep(1.2)
            try:
              #  driver.implicitly_wait(10)
                WebDriverWait(driver,10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "a[href='{}']".format(href)))).click()
                time.sleep(1)
               # driver.find_element_by_css_selector("a[href='{}']".format(href)).click()
              #  time.sleep(1.2)
            except:
                pass
            finally:
                if(start-10 > 0):
                    xNo+=10

    for m in range(start%10):
        href = "#entrno={}&mallid={}&searchname=&filtername=NAME&PageNumber=".format(str(sellerID), seller) + str(xNo)
        #time.sleep(1.2)
        try:
            #driver.implicitly_wait(10)
            WebDriverWait(driver,20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "a[href='{}']".format(href)))).click()
            time.sleep(1)
          #  driver.find_element_by_css_selector("a[href='{}']".format(href)).click()
          #  time.sleep(1.2)
        except:
            pass
        finally:
                xNo+=1

    for j in range(1,max):
        href = "#entrno={}&mallid={}&searchname=&filtername=NAME&PageNumber=".format(str(sellerID), seller) + str(xNo)
        try:
            time.sleep(1)
            lst = WebDriverWait(driver,10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'strong.used-class-lowest')))
            for i in range(len(lst)):
                try:
                    time.sleep(0.5)
                    lst = WebDriverWait(driver,10).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, 'strong.used-class-lowest')))
                    print("lst = {}".format(len(lst)))
                    lst[i].click()
                    title = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.TAG_NAME, 'h2.gd_name'))).text
                    price = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'em.yes_m'))).text
                    auth = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'span.gd_auth'))).text
                    publisher = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'span.gd_pub'))).text
                    date = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'span.gd_date'))).text
                    ownPrice = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div/div[4]/div[2]/div[2]/div[1]/div[1]/table/tbody/tr[1]/td'))).text
                    price = price.replace(",", "")
                    ownPrice = ownPrice.replace(",", "")
                    ownPrice = ownPrice.replace("원", "")
                    Yes24WriteExcel(count, title, price, auth, publisher, date, count, ownPrice)    
                    print(i)
                finally:
                    driver.back()
                    count+=1
                    pass
        except:
            pass

        scrollDown(driver)

        try:
            driver.implicitly_wait(10)
            a = WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[href='{}']".format(href))))
            a.click()
            time.sleep(1)
            xNo+=1
            #driver.find_element_by_css_selector("a[href='{}']".format(href)).click()
        except:
            xNo-=1
            pass
            
def scrollUp(driver):
    start = datetime.datetime.now() 
    end = start + datetime.timedelta(seconds=1)
    while True:
        driver.execute_script('window.scrollTo(document.body.scrollHeight, 0);')
        if datetime.datetime.now() > end:
            break


def scrollDown(driver):
    start = datetime.datetime.now() 
    end = start + datetime.timedelta(seconds=1)
    while True:
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        if datetime.datetime.now() > end:
            break


def Yes24WriteExcel(isbn, title, price, auth, publisher, date, count, ownPrice):
    excel_name = str(entry_excel.get()) + ".xlsx"
    wb = load_workbook(excel_name)
    ws = wb.active
    lst = ["", isbn, 1111, "A", "A", "N", "1", "", int(price)-50, title, "", "", "품질보장~!! (미사용 ★ 출판사에서 직접구매한 새★책 ^^)  ■  >□<", "", "1", "1",publisher, ownPrice]
    ws.append(lst)
    wb.save(excel_name)
    print("ok")
    wb.close()

main()


